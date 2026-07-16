"""
XLSX-index permit source.

For jurisdictions that publish permits as a page of downloadable Excel files
(one per period) instead of a GIS/API feed — e.g. City of Boca Raton's
"Applied/Issued Permits" page, which links monthly `AppliedPermits_YYYY-MM-DD.xlsx`
and `IssuedPermits_*.xlsx` files.

Flow:
  1. GET the index page, extract the permit .xlsx links (anchor text + href).
  2. Pick the files whose period overlaps the requested window, newest first,
     capped for the request budget.
  3. Download each and parse rows with openpyxl (read-only).
  4. Map columns to canonical permit fields by header name.

Never raises — returns (rows, diag). Standard library + requests + openpyxl.
"""

import io
import re
from datetime import datetime

import requests

UA = "CollaborativeConcepts-PermitLeads/1.0 (+https://collaborativeconceptsfl.com)"
TIMEOUT = 12
MAX_FILES = 4          # cap downloads per request (each xlsx can be large)
DEFAULT_DEADLINE = 20  # wall-clock seconds


# Header synonyms -> canonical field (normalized: uppercase alnum only)
_HEADER_SYNONYMS = {
    "permit_number": ["PERMITNUMBER", "PERMITNO", "PERMITNUM", "PERMIT", "PERMITID",
                      "RECORDNUMBER", "CASENUMBER", "APPLICATIONNUMBER", "NUMBER"],
    "description":   ["DESCRIPTION", "WORKDESCRIPTION", "SCOPEOFWORK", "SCOPE",
                      "PROJECTDESCRIPTION", "PROJECTNAME", "JOBDESCRIPTION", "WORK",
                      "PERMITDESCRIPTION", "DESCR", "TYPEOFWORK"],
    "address":       ["ADDRESS", "SITEADDRESS", "JOBADDRESS", "PROPERTYADDRESS",
                      "FULLADDRESS", "LOCATION", "SITEADDR", "PROJECTADDRESS"],
    "issue_date":    ["ISSUEDATE", "ISSUEDDATE", "DATEISSUED", "ISSUED",
                      "PERMITISSUEDATE"],
    "applied_date":  ["APPLIEDDATE", "APPLICATIONDATE", "DATEAPPLIED", "APPLYDATE",
                      "SUBMITTEDDATE", "SUBMITDATE", "CREATEDDATE", "DATE"],
    "status":        ["STATUS", "PERMITSTATUS", "STATUSDESCRIPTION", "CURRENTSTATUS"],
    "type":          ["PERMITTYPE", "TYPE", "WORKTYPE", "PERMITTYPEDESCRIPTION",
                      "CATEGORY", "CLASS", "APPLICATIONTYPE", "SUBTYPE"],
    "contractor":    ["CONTRACTOR", "CONTRACTORNAME", "COMPANY", "COMPANYNAME",
                      "BUSINESSNAME", "GENERALCONTRACTOR", "CONTRACTORCOMPANY"],
    "value":         ["VALUATION", "JOBVALUE", "ESTIMATEDVALUE", "CONSTRUCTIONVALUE",
                      "VALUE", "JOBCOST", "ESTIMATEDCOST", "CONSTRUCTIONCOST",
                      "DECLAREDVALUE", "TOTALVALUE", "PERMITVALUE"],
    "owner":         ["OWNER", "OWNERNAME", "PROPERTYOWNER", "OWNERSNAME"],
    "city":          ["CITY", "SITECITY"],
    "zip":           ["ZIP", "ZIPCODE", "POSTALCODE", "ZIP5"],
}
_DATE_IN_NAME = re.compile(r"(20\d{2})[-_/](\d{1,2})(?:[-_/](\d{1,2}))?")


def _norm(s):
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def _session():
    s = requests.Session()
    s.headers.update({"User-Agent": UA, "Accept": "text/html,*/*"})
    return s


def extract_xlsx_links(html, base_url):
    """Return [{url, text, date}] for permit .xlsx anchors on the index page.
    Handles CivicPlus /DocumentCenter/View/<id>/<name>-xlsx links whose href
    may not literally end in .xlsx (the filename lives in the anchor text)."""
    out = []
    for m in re.finditer(r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>',
                         html or "", re.I | re.S):
        href, text = m.group(1), re.sub(r"<[^>]+>", "", m.group(2))
        text = text.strip()
        blob = f"{href} {text}"
        if not re.search(r"permit", blob, re.I):
            continue
        if not (re.search(r"\.xlsx?\b", blob, re.I) or
                re.search(r"documentcenter/view", href, re.I)):
            continue
        # absolutize
        url = href
        if url.startswith("//"):
            url = "https:" + url
        elif url.startswith("/"):
            m2 = re.match(r"^(https?://[^/]+)", base_url)
            url = (m2.group(1) if m2 else "") + url
        elif not url.startswith("http"):
            url = base_url.rstrip("/") + "/" + url
        dm = _DATE_IN_NAME.search(text) or _DATE_IN_NAME.search(href)
        d = None
        if dm:
            y, mo, day = int(dm.group(1)), int(dm.group(2) or 1), int(dm.group(3) or 1)
            try:
                d = datetime(y, mo, day)
            except ValueError:
                d = None
        kind = "issued" if re.search(r"issued", blob, re.I) else "applied"
        out.append({"url": url, "text": text or href, "date": d, "kind": kind})
    return out


def _pick_files(links, start_dt, end_dt, max_files=MAX_FILES):
    """Files whose period overlaps [start,end], newest first, capped. Files
    with no parseable date are kept as a low-priority fallback."""
    start_naive = start_dt.replace(tzinfo=None) if start_dt.tzinfo else start_dt
    end_naive = end_dt.replace(tzinfo=None) if end_dt.tzinfo else end_dt
    dated = [l for l in links if l["date"]]
    undated = [l for l in links if not l["date"]]
    # a monthly file dated M covers roughly [M, M+~1mo]; keep files from just
    # before the window through the end.
    keep = [l for l in dated if start_naive.replace(day=1) <= l["date"] <= end_naive
            or (l["date"] <= start_naive and (start_naive - l["date"]).days <= 31)]
    keep.sort(key=lambda l: l["date"], reverse=True)
    if not keep:
        keep = sorted(dated, key=lambda l: l["date"], reverse=True)[:max_files]
    return (keep or undated)[:max_files]


def _map_headers(headers):
    by_norm = {}
    for i, h in enumerate(headers):
        n = _norm(h)
        if n and n not in by_norm:
            by_norm[n] = i
    mapping = {}
    for canon, syns in _HEADER_SYNONYMS.items():
        for syn in syns:
            if syn in by_norm:
                mapping[canon] = by_norm[syn]
                break
        if canon not in mapping:
            for n, idx in by_norm.items():
                if any((len(s) > 3 and (s in n or n in s)) for s in syns):
                    mapping[canon] = idx
                    break
    return mapping


def _cell_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() or None


def parse_xlsx(content, max_rows=5000):
    """Parse xlsx bytes -> (list-of-row-dicts keyed by canonical field, headers).
    Uses the first row as headers. Returns ([], []) on any failure."""
    try:
        import openpyxl
    except ImportError:
        return [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return [], []
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = None
        for r in rows_iter:
            if r and any(c is not None and str(c).strip() for c in r):
                headers = [(_cell_str(c) or "") for c in r]
                break
        if not headers:
            return [], []
        mapping = _map_headers(headers)
        out = []
        for r in rows_iter:
            if not r or all(c is None or str(c).strip() == "" for c in r):
                continue
            rec = {}
            for canon, idx in mapping.items():
                if idx < len(r):
                    rec[canon] = _cell_str(r[idx])
            if any(rec.get(k) for k in ("permit_number", "address", "description")):
                out.append(rec)
            if len(out) >= max_rows:
                break
        return out, headers
    finally:
        try:
            wb.close()
        except Exception:
            pass


def fetch_xlsx_index(index_url, start_dt, end_dt, deadline=DEFAULT_DEADLINE,
                     max_files=MAX_FILES):
    """Fetch + parse the permit spreadsheets overlapping the window.
    Returns (rows, diag). rows carry canonical keys (permit_number, description,
    address, issue_date/applied_date, status, type, contractor, value, ...).
    Never raises."""
    import time as _time
    started = _time.time()
    diag = {"index_url": index_url, "files": [], "steps": []}
    sess = _session()
    try:
        r = sess.get(index_url, timeout=TIMEOUT)
        diag["steps"].append(f"GET index: {r.status_code}")
        if r.status_code != 200:
            diag["error"] = f"index HTTP {r.status_code}"
            return [], diag
        links = extract_xlsx_links(r.text, index_url)
        diag["links_found"] = len(links)
        if not links:
            diag["error"] = "no permit .xlsx links found on index page"
            return [], diag
        chosen = _pick_files(links, start_dt, end_dt, max_files)
        diag["files_chosen"] = [l["text"] for l in chosen]
        rows = []
        for l in chosen:
            if (_time.time() - started) > deadline:
                diag["steps"].append("stopped: deadline")
                break
            try:
                fr = sess.get(l["url"], timeout=TIMEOUT)
                if fr.status_code != 200:
                    diag["files"].append({"file": l["text"], "error": f"HTTP {fr.status_code}"})
                    continue
                parsed, headers = parse_xlsx(fr.content)
                diag["files"].append({"file": l["text"], "rows": len(parsed),
                                      "headers": headers[:20] if headers else []})
                for rec in parsed:
                    rec.setdefault("applied_date", None)
                    # tag the file's kind so issue/applied date is sensible
                    if l["kind"] == "issued" and not rec.get("issue_date") and rec.get("applied_date"):
                        pass
                    rows.append(rec)
            except Exception as e:
                diag["files"].append({"file": l["text"], "error": str(e)[:120]})
        diag["parsed"] = len(rows)
        if not rows and not diag.get("error"):
            diag["error"] = "downloaded files but parsed no permit rows"
        return rows, diag
    except requests.Timeout:
        diag["error"] = "timeout contacting index/site"
        return [], diag
    except Exception as e:
        diag["error"] = f"{type(e).__name__}: {str(e)[:160]}"
        return [], diag
